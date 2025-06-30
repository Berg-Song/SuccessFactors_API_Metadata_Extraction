import openpyxl
import json
import os
from openpyxl.styles import Border, Side
import requests

EMPLOYEE_ID = "Berg01"
POSITION = "10023800"
EVENT_REASON = "H-001"
HIRE_DATE = "/Date(1743033600000)/"
TEST_API_SERVER = "apiDemopreview.sapsf.com"
API_ENDPOINT = f"https://{TEST_API_SERVER}/odata/v2/upsert?$format=json&$purgeType=full"
RELATED_PERSONIDEXTERNAL = "Berg01_01"
BACKGROUND_ID="0"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "New Hire API DocumentV1.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "New Hire API Post Preview.xlsx")

MAX_SHEETNAME_LEN = 31

def transform_json(json_str, entity_name=None):
    # Try to parse JSON (handle single/double quotes)
    try:
        data = json.loads(json_str)
    except Exception:
        data = json.loads(json_str.replace("'", '"'))
    # Replace values as required
    for key in list(data.keys()):
        if key in ["userId", "personIdExternal", "username", "PaymentInformationV3_worker", "worker"]:
            data[key] = EMPLOYEE_ID
        elif key == "emailAddress":
            data[key] = f"{EMPLOYEE_ID}@dummy.com"
        elif key == "position":
            data[key] = POSITION
        elif key == "backgroundElementId":
            data[key] = BACKGROUND_ID
        elif key == "eventReason":
            data[key] = EVENT_REASON
        elif key == "relatedPersonIdExternal":
            data[key] = RELATED_PERSONIDEXTERNAL
        elif key in ["startDate", "PaymentInformationV3_effectiveStartDate", "effectiveStartDate", "payDate"]:
            # Only replace if entity_name does NOT start with "Background"
            if not (entity_name and str(entity_name).startswith("Background")):
                data[key] = HIRE_DATE
        elif key == "__metadata" and isinstance(data[key], dict):
            if "uri" in data[key]:
                data[key]["uri"] = data[key]["uri"].replace("apiDemo.successfactors.com", TEST_API_SERVER)
    return data

def remove_blank_values(d):
    """Remove keys where value is None or blank string."""
    return {k: v for k, v in d.items() if v not in [None, ""]}

def get_valid_sheet_name(entity_name):
    sheet_name = str(entity_name)
    if sheet_name.startswith("Background_"):
        # Always remove the prefix, regardless of length
        sheet_name = sheet_name[len("Background_"):]
    # Truncate to 31 chars if still too long
    if len(sheet_name) > MAX_SHEETNAME_LEN:
        sheet_name = sheet_name[:MAX_SHEETNAME_LEN]
    return sheet_name

def update_api_templates():
    INTEGRATION_FILE = os.path.join(SCRIPT_DIR, "DiDi SF New Hire IntegrationV1.xlsx")
    integration_wb = openpyxl.load_workbook(INTEGRATION_FILE)
    table_ws = integration_wb["SF Master Table List"]
    template_ws = integration_wb["API Template"]
    sf_master_ws = integration_wb["SF Master Data Dictionary"]

    # Get headers and their indices
    headers = [cell.value for cell in table_ws[1]]
    entity_idx = headers.index("Entity")
    api_name_idx = headers.index("API Name")
    data_flow_idx = headers.index("Data Flow")
    trigger_point_idx = headers.index("Trigger Point")
    test_api_idx = headers.index("Test API Endpoint")
    pro_api_idx = headers.index("Pro API Endpoint")
    http_method_idx = headers.index("HTTP Method")
    sample_upsert_idx = headers.index("Sample Upsert")
    sample_response_idx = headers.index("Sample Response")

    sf_master_headers = [cell.value for cell in sf_master_ws[1]]
    sf_entity_col_idx = sf_master_headers.index("Entity")  # 0-based

    # Collect all entity rows (skip header)
    entities = []
    for row in table_ws.iter_rows(min_row=2, values_only=True):
        entity = row[entity_idx]
        if entity and str(entity).strip():
            entities.append({
                "Entity": entity,
                "API Name": row[api_name_idx],
                "Data Flow": row[data_flow_idx],
                "Trigger Point": row[trigger_point_idx],
                "Test API Endpoint": row[test_api_idx],
                "Pro API Endpoint": row[pro_api_idx],
                "HTTP Method": row[http_method_idx],
                "Sample Upsert": row[sample_upsert_idx],
                "Sample Response": row[sample_response_idx]
            })

    # Remove old entity sheets if they exist
    # for entity in entities: ...

    # Duplicate, populate, and copy B–I for each entity
    for entity in entities:
        sheet_name = get_valid_sheet_name(entity["Entity"])
        # Delete the sheet if it already exists
        if sheet_name in integration_wb.sheetnames:
            std = integration_wb[sheet_name]
            integration_wb.remove(std)
        new_ws = integration_wb.copy_worksheet(template_ws)
        new_ws.title = sheet_name

        # Populate variables
        new_ws["A1"].value = str(entity["API Name"]) if entity["API Name"] is not None else ""
        new_ws["B2"].value = entity["Data Flow"]
        new_ws["B3"].value = entity["Trigger Point"]
        new_ws["B4"].value = entity["Entity"]
        new_ws["B5"].value = entity["Test API Endpoint"]
        new_ws["B6"].value = entity["Pro API Endpoint"]
        new_ws["B7"].value = entity["HTTP Method"]
        new_ws["B11"].value = entity["Sample Upsert"]
        new_ws["K11"].value = entity["Sample Response"]

        # Copy B–I from SF Master Data Dictionary to B14–I* in the entity sheet
        matching_rows = []
        for row in sf_master_ws.iter_rows(min_row=2, values_only=True):
            if get_valid_sheet_name(row[sf_entity_col_idx]) == sheet_name:
                matching_rows.append(row)
        for i, row in enumerate(matching_rows):
            for col_offset in range(1, 9):  # B=2, ..., I=9 (1-based)
                value = row[col_offset]
                new_ws.cell(row=14 + i, column=col_offset + 1, value=value)

    integration_wb.save(INTEGRATION_FILE)
    print(f"API Template sheets updated and B–I copied for all entities in {INTEGRATION_FILE}")

def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb["API Entity"]
    headers = [cell.value for cell in ws[1]]
    api_sample_idx = headers.index("Api sample upsert")
    entity_idx = headers.index("Entity")

    # Prepare output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "API Post Preview"
    out_ws.append(["Entity", "API Endpoint", "Body", "Valid Body"])

    for row in ws.iter_rows(min_row=2, values_only=True):
        entity = row[entity_idx]
        json_str = row[api_sample_idx]
        if not json_str:
            continue
        try:
            body = transform_json(json_str, entity)
            valid_body = remove_blank_values(body)
            # Special handling for User entity __metadata.uri
            if entity == "User" and "__metadata" in valid_body and isinstance(valid_body["__metadata"], dict):
                uri = valid_body["__metadata"].get("uri", "")
                # Only append if not already present
                if not uri.endswith(f"('{EMPLOYEE_ID}')"):
                    if uri.endswith("User"):
                        uri += f"('{EMPLOYEE_ID}')"
                        valid_body["__metadata"]["uri"] = uri
        except Exception as e:
            print(f"Error parsing JSON for entity {entity}: {e}")
            continue
        out_ws.append([
            entity,
            API_ENDPOINT,
            json.dumps(body, ensure_ascii=False),
            json.dumps(valid_body, ensure_ascii=False)
        ])
        print(f"Entity: {entity}\nEndpoint: {API_ENDPOINT}\nBody: {json.dumps(body, ensure_ascii=False)}\nValid Body: {json.dumps(valid_body, ensure_ascii=False)}\n")

    out_wb.save(OUTPUT_FILE)
    print(f"Preview exported to {OUTPUT_FILE}")

    INTEGRATION_FILE = os.path.join(SCRIPT_DIR, "DiDi SF New Hire IntegrationV1.xlsx")
    integration_wb = openpyxl.load_workbook(INTEGRATION_FILE, data_only=True)
    integration_ws = integration_wb["SF Master Table List"]

    # Build a lookup from API Post Preview for Entity -> Valid Body
    preview_wb = openpyxl.load_workbook(OUTPUT_FILE)
    preview_ws = preview_wb["API Post Preview"]
    preview_headers = [cell.value for cell in preview_ws[1]]
    entity_idx_preview = preview_headers.index("Entity")
    valid_body_idx = preview_headers.index("Valid Body")
    entity_to_valid_body = {}
    for row in preview_ws.iter_rows(min_row=2, values_only=True):
        entity = row[entity_idx_preview]
        valid_body = row[valid_body_idx]
        entity_to_valid_body[str(entity)] = valid_body

    # Find column indices in integration sheet
    integration_headers = [cell.value for cell in integration_ws[1]]
    entity_idx_integration = integration_headers.index("Entity")
    sample_upsert_idx = integration_headers.index("Sample Upsert")

    # Populate Sample Upsert with Valid Body where Entity matches
    for row in integration_ws.iter_rows(min_row=2):
        entity = str(row[entity_idx_integration].value)
        if entity in entity_to_valid_body:
            integration_ws.cell(row=row[0].row, column=sample_upsert_idx + 1, value=entity_to_valid_body[entity])

    # Copy and transform "Upsert API Field Attribute" to "SF Master Data Dictionary"
    upsert_ws = wb["Upsert API Field Attribute"]
    sf_master_ws = integration_wb["SF Master Data Dictionary"]

    # Clear existing data except header in SF Master Data Dictionary
    sf_master_ws.delete_rows(2, sf_master_ws.max_row - 1)

    # Get headers and indices
    upsert_headers = [cell.value for cell in upsert_ws[1]]
    field_idx = upsert_headers.index("Field")
    sample_value_idx = upsert_headers.index("Sample value")

    # Only append headers if not present
    if all(cell.value is None for cell in sf_master_ws[1]):
        sf_master_ws.append(upsert_headers)

    for row in upsert_ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        field = str(row[field_idx]) if row[field_idx] else ""
        sample_value = row[sample_value_idx]
        # Apply transformation based on Field
        if field in ["userId", "personIdExternal", "username"]:
            row[sample_value_idx] = EMPLOYEE_ID
        elif field == "emailAddress":
            row[sample_value_idx] = f"{EMPLOYEE_ID}@dummy.com"
        elif field == "__metadata" and sample_value and "apiDemo.successfactors.com" in str(sample_value):
            row[sample_value_idx] = str(sample_value).replace("apiDemo.successfactors.com", TEST_API_SERVER)
        sf_master_ws.append(row)

    # Add border to all cells (including blank)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sf_master_ws.iter_rows(min_row=1, max_row=sf_master_ws.max_row, max_col=sf_master_ws.max_column):
        for cell in row:
            cell.border = border

    integration_wb.save(INTEGRATION_FILE)
    print(f'Upsert API Field Attribute copied and Sample Value transformed in "SF Master Data Dictionary" of {INTEGRATION_FILE}')

def post_valid_bodies_and_export_response():
    preview_wb = openpyxl.load_workbook(OUTPUT_FILE)
    preview_ws = preview_wb["API Post Preview"]
    headers = [cell.value for cell in preview_ws[1]]
    valid_body_idx = headers.index("Valid Body")
    # Add a new column for API Response if not already present
    if "API Response" not in headers:
        preview_ws.cell(row=1, column=len(headers) + 1, value="API Response")
        response_col = len(headers) + 1
    else:
        response_col = headers.index("API Response") + 1

    # Open integration workbook and get SF Master Table List
    INTEGRATION_FILE = os.path.join(SCRIPT_DIR, "DiDi SF New Hire IntegrationV1.xlsx")
    integration_wb = openpyxl.load_workbook(INTEGRATION_FILE, data_only=True)
    integration_ws = integration_wb["SF Master Table List"]
    integration_headers = [cell.value for cell in integration_ws[1]]
    entity_idx_integration = integration_headers.index("Entity")
    sample_upsert_idx = integration_headers.index("Sample Upsert")
    sample_response_idx = integration_headers.index("Sample Response")

    for i, row in enumerate(preview_ws.iter_rows(min_row=2), start=2):
        valid_body = row[valid_body_idx].value
        entity = row[0].value
        if not valid_body or valid_body.strip() in ["", "{}", "null"]:
            preview_ws.cell(row=i, column=response_col, value="Skipped: Blank body")
            print(f"Row {i}: Skipped (blank body)")
            continue
        try:
            response = requests.post(
                API_ENDPOINT,
                data=valid_body.encode("utf-8"),
                headers={"Content-Type": "application/json"},
                auth=("mdmapi@xiaojuscieT1", "sfmdmapi123"),
                timeout=30
            )
            api_response = response.text
            preview_ws.cell(row=i, column=response_col, value=api_response)
            print(f"Row {i}: Response: {response.status_code} {api_response}")
        except Exception as e:
            api_response = f"Error: {e}"
            preview_ws.cell(row=i, column=response_col, value=api_response)
            print(f"Row {i}: Error: {e}")

        # --- Copy Valid Body and API Response to SF Master Table List ---
        for int_row in integration_ws.iter_rows(min_row=2):
            if str(int_row[entity_idx_integration].value) == str(entity):
                integration_ws.cell(row=int_row[0].row, column=sample_upsert_idx + 1, value=valid_body)
                integration_ws.cell(row=int_row[0].row, column=sample_response_idx + 1, value=api_response)
                break

    preview_wb.save(OUTPUT_FILE)
    integration_wb.save(INTEGRATION_FILE)
    print(f"API responses exported to {OUTPUT_FILE} and copied to SF Master Table List")

# Call this at the end of your main()
if __name__ == "__main__":
    main()
    post_valid_bodies_and_export_response()
    update_api_templates()
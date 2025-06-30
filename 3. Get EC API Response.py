import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import json

API_SERVER = "apiDemo.successfactors.com"
USERNAME = "Berg@CompanyId"
PASSWORD = "Berg"
EmployeeId = "Berg001"  # Replace with actual Employee ID, it's used for filtering API results
ENTITY_SETS = [
    "User", "PerPerson", "EmpEmployment", "EmpJob", "PerPersonal","PerGlobalInfoMEX",
    "EmpJobRelationships", "EmpCompensation","EmpPayCompRecurring","EmpPayCompNonRecurring","EmpWorkPermit",  "PerNationalId", "PerEmail",
    "PerPhone", "PerPersonRelationship", "PerAddressDEFLT", "PerEmergencyContacts","PaymentInformationV3",
    "PaymentInformationDetailV3", "Background_OutsideWorkExperience", "Background_Education",
    "Background_Certificates", "Background_Languages"
]
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "SF New Hire API UpsertV1.xlsx")

def get_filter(entity):
    if entity.startswith("Emp"):
        return "userId"
    elif entity.startswith("Per"):
        return "personIdExternal"
    elif entity.startswith("Background"):
        return "userId"
    elif entity.startswith("cust"):
        return "externalCode"
    elif entity.startswith("EmpCostDistribution"):
        return "usersSysId"
    elif entity.startswith("PaymentInformationV3"):
        return "worker"
    elif entity.startswith("PaymentInformationDetailV3"):
        return "PaymentInformationV3_worker"
    else:
        return "userId"

def clean_metadata_url(url):
    # Remove content between ( and )
    return re.sub(r"\(.*?\)", "", url)

def clean_json(data):
    if isinstance(data, dict):
        cleaned = {}
        for k, v in data.items():
            # Remove keys as specified
            if k.startswith("created") or k.startswith("lastModified") or "Nav" in k:
                continue
            # Skip key-value if value is a dict, except for __metadata
            if isinstance(v, dict) and k != "__metadata":
                continue
            # Clean __metadata.url and __metadata.uri recursively
            if k == "__metadata" and isinstance(v, dict):
                if "url" in v:
                    v["url"] = clean_metadata_url(v["url"])
                if "uri" in v:
                    v["uri"] = clean_metadata_url(v["uri"])
            cleaned[k] = clean_json(v)
        return cleaned
    elif isinstance(data, list):
        return [clean_json(item) for item in data]
    elif data is None:
        return ""
    else:
        return data

def autofit_and_style(ws):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Autofit columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    # Autofit rows
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    # Green header
    for cell in ws[1]:
        cell.fill = green_fill

def main():
    api_entity_rows = []
    field_attr_rows = []
    for entity in ENTITY_SETS:
        filter_field = get_filter(entity)
        if entity in ["PerNationalId", "PerEmail", "PerPhone"]:
            endpoint = f"https://{API_SERVER}/odata/v2/{entity}?$format=json&$filter={filter_field} eq '{EmployeeId}' and isPrimary eq true"
        else:
            endpoint = f"https://{API_SERVER}/odata/v2/{entity}?$format=json&$filter={filter_field} eq '{EmployeeId}'"
        try:
            resp = requests.get(endpoint, auth=HTTPBasicAuth(USERNAME, PASSWORD), verify=True)
            print(f"Response for {entity}: {resp.text}")
            resp.raise_for_status()
            d = resp.json()
            result = d.get('d', {}).get('results') or d.get('d', {}).get('result', [])
            # If result is blank, try again without filter
            if (isinstance(result, list) and not result) or (isinstance(result, dict) and not result):
                endpoint = f"https://{API_SERVER}/odata/v2/{entity}?$format=json&$top=1"
                resp = requests.get(endpoint, auth=HTTPBasicAuth(USERNAME, PASSWORD), verify=True)
                print(f"Fallback response for {entity}: {resp.text}")
                resp.raise_for_status()
                d = resp.json()
                result = d.get('d', {}).get('results') or d.get('d', {}).get('result', [])
            if isinstance(result, list):
                result = result[0] if result else {}
            elif isinstance(result, dict):
                result = result
            else:
                result = {}
            if "__metadata" in result and "url" in result["__metadata"]:
                result["__metadata"]["url"] = clean_metadata_url(result["__metadata"]["url"])
            cleaned = clean_json(result)
            # Sheet 1
            api_entity_rows.append([entity, endpoint, json.dumps(cleaned, ensure_ascii=False)])
            # Sheet 2
            for k, v in cleaned.items():
                if isinstance(v, (dict, list)):
                    v = str(v)
                field_attr_rows.append([entity, k, v])
        except Exception as e:
            api_entity_rows.append([entity, endpoint, f"Error: {e}"])
            continue

    # Write to Excel
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "API Entity"
    ws2 = wb.create_sheet("Upsert API Field Attribute")

    ws1.append(["Entity", "API Endpoint", "API Sample Upsert"])
    for row in api_entity_rows:
        ws1.append(row)
    autofit_and_style(ws1)

    ws2.append(["Entity", "Field", "Sample Value"])
    for row in field_attr_rows:
        ws2.append(row)
    autofit_and_style(ws2)

    wb.save(EXCEL_FILE)
    print(f"Done. Output: {os.path.abspath(EXCEL_FILE)}")

if __name__ == "__main__":
    main()
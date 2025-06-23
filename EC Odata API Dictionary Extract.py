import requests
from requests.auth import HTTPBasicAuth
import xml.etree.ElementTree as ET
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import os

# API variables
API_SERVER = "api4.successfactors.com"
ENTITY_SETS = [
    "User", "PerPerson", "EmpEmployment", "EmpJob", "PerPersonal", "EmpEmploymentTermination",
    "EmpJobRelationships", "PerEmail", "EmpCompensation", "PerNationalId", "EmpWorkPermit",
    "PerPhone", "PerPersonRelationship", "PerAddressDEFLT", "PerEmergencyContacts",
    "PaymentInformationDetailV3", "Background_OutsideWorkExperience", "Background_Education",
    "Background_Certificates", "Background_Languages", "TrendData_SysOverallPerformance"
]
USERNAME = "Berg@CompanyId"
PASSWORD = "Berg"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_OUTPUT_PATH = os.path.join(SCRIPT_DIR, "EC_APIField_Metadata.xlsx") # <-- moved here

# Download and parse metadata for each entity set
metadata_trees = {}
for entity in ENTITY_SETS:
    url = f"https://{API_SERVER}/odata/v2/{entity}/$metadata"
    print(f"Fetching metadata for {entity}...")
    response = requests.get(url, auth=HTTPBasicAuth(USERNAME, PASSWORD))
    if response.status_code == 200:
        metadata_trees[entity] = ET.ElementTree(ET.fromstring(response.content))
    else:
        print(f"Failed to fetch metadata for {entity}: {response.status_code}")

ns = {
    'edmx': 'http://schemas.microsoft.com/ado/2007/06/edmx',
    'm': 'http://schemas.microsoft.com/ado/2007/08/dataservices/metadata',
    '': 'http://schemas.microsoft.com/ado/2008/09/edm',
    'sap': 'http://www.successfactors.com/edm/sap'
}

def get_text(node, tag):
    el = node.find(tag, ns)
    return el.text if el is not None else "Null"

def get_sap_tags(doc):
    tags = doc.find('.//sap:tagcollection', ns)
    if tags is not None:
        return ', '.join([t.text for t in tags.findall('sap:tag', ns)])
    return "Null"

def get_attr_value(node, attr):
    if attr in node.attrib:
        return node.attrib[attr]
    if 'sap:' + attr in node.attrib:
        return node.attrib['sap:' + attr]
    sap_ns = '{http://www.successfactors.com/edm/sap}' + attr
    if sap_ns in node.attrib:
        return node.attrib[sap_ns]
    return "Null"

def clean_attr(attr):
    return re.sub(r'^\{.*\}', '', attr).replace('sap:', '')

# 1. EC Entity Sheet (collect from all metadata trees)
entities = []
entityset_cols = ['Name', 'label', 'creatable', 'updatable', 'upsertable', 'deletable']
for tree in metadata_trees.values():
    root = tree.getroot()
    for schema in root.findall('.//{http://schemas.microsoft.com/ado/2008/09/edm}Schema'):
        if schema.attrib.get('Namespace') == 'SFODataSet':
            for es in schema.findall('.//EntitySet', ns):
                doc = es.find('Documentation', ns)
                summary = get_text(doc, 'Summary') if doc is not None else "Null"
                longdesc = get_text(doc, 'LongDescription') if doc is not None else "Null"
                tags = get_sap_tags(doc) if doc is not None else "Null"
                row = {}
                for col in entityset_cols:
                    if col == 'Name':
                        row['Name'] = es.attrib.get('Name', 'Null')
                    else:
                        row[col] = get_attr_value(es, col)
                row['Summary'] = summary
                row['LongDescription'] = longdesc
                row['Sap Tagcollection'] = tags
                entities.append(row)

df_entities = pd.DataFrame(entities, columns=entityset_cols + ['Summary', 'LongDescription', 'Sap Tagcollection'])

# 2. EC Data API Dictionary Sheet (collect from all metadata trees)
rows = []
all_attrs = set()

for tree in metadata_trees.values():
    root = tree.getroot()
    for schema in root.findall('.//{http://schemas.microsoft.com/ado/2008/09/edm}Schema'):
        if schema.attrib.get('Namespace') == 'SFOData':
            for et in schema.findall('EntityType', ns):
                for prop in et.findall('Property', ns):
                    all_attrs.update([clean_attr(a) for a in prop.attrib.keys()])
                for nav in et.findall('NavigationProperty', ns):
                    all_attrs.update([clean_attr(a) for a in nav.attrib.keys()])

all_attrs = sorted(all_attrs)
all_attrs += ["Key", "Entity", "NavigationField"]

for tree in metadata_trees.values():
    root = tree.getroot()
    for schema in root.findall('.//{http://schemas.microsoft.com/ado/2008/09/edm}Schema'):
        if schema.attrib.get('Namespace') == 'SFOData':
            for et in schema.findall('EntityType', ns):
                entity_name = et.attrib.get('Name', 'Null')
                key_names = set()
                key = et.find('Key', ns)
                if key is not None:
                    key_names = {pr.attrib.get('Name') for pr in key.findall('PropertyRef', ns)}
                for prop in et.findall('Property', ns):
                    row = {}
                    for attr in all_attrs:
                        if attr in ["Key", "Entity", "NavigationField"]:
                            continue
                        row[attr] = get_attr_value(prop, attr)
                    row["Key"] = "true" if prop.attrib.get("Name") in key_names else "false"
                    row["Entity"] = entity_name
                    row["NavigationField"] = "false"
                    rows.append(row)
                for nav in et.findall('NavigationProperty', ns):
                    row = {}
                    for attr in all_attrs:
                        if attr in ["Key", "Entity", "NavigationField"]:
                            continue
                        row[attr] = get_attr_value(nav, attr)
                    row["Key"] = "false"
                    row["Entity"] = entity_name
                    row["NavigationField"] = "true"
                    rows.append(row)

df_dict = pd.DataFrame(rows, columns=all_attrs)

# 3. Simple EC Data API Dictionary Sheet
simple_cols = [
    "Entity", "Name", "Key", "required", "picklist", "MaxLength", "NavigationField",
    "visible", "filterable", "sortable", "upsertable"
]
df_simple = df_dict[simple_cols].copy()

# Sorting: Entity (asc), Name (asc), Key (desc), required (desc)
def sort_key(row):
    # Key and required: "true" > "false"
    return (
        row['Entity'] if pd.notnull(row['Entity']) else "",
        row['Name'] if pd.notnull(row['Name']) else "",
        1 if str(row['Key']).lower() == "true" else 0,
        1 if str(row['required']).lower() == "true" else 0
    )

df_simple = df_simple.sort_values(
    by=["Entity", "Name", "Key", "required"],
    ascending=[True, True, False, False],
    key=lambda col: col.map(lambda x: 1 if str(x).lower() == "true" else 0) if col.name in ["Key", "required"] else col
).reset_index(drop=True)

# 4. Write to Excel
with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
    df_entities.to_excel(writer, sheet_name="EC Entity", index=False)
    df_dict.to_excel(writer, sheet_name="EC Data API Dictionary", index=False)
    df_simple.to_excel(writer, sheet_name="Simple EC Data API Dictionary", index=False)

# 5. Formatting with openpyxl
wb = load_workbook(EXCEL_OUTPUT_PATH)
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
thin = Side(border_style="thin", color="000000")

for ws in wb.worksheets:
    # Autofit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    # Add border and green fill for header
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.row == 1:
                cell.fill = green_fill
            if cell.value is not None:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb.save(EXCEL_OUTPUT_PATH)
print(f"Excel file '{EXCEL_OUTPUT_PATH}' created successfully.")
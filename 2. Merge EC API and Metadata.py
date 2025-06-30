import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import json

def capitalize_headers(ws):
    for cell in ws[1]:
        if cell.value:
            cell.value = str(cell.value).capitalize()

def autofit_and_style(ws):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Autofit columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            cell.border = border  # Always apply border
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    # Set a reasonable row height for all rows with value
    for row in ws.iter_rows():
        has_value = any(cell.value for cell in row)
        if has_value:
            ws.row_dimensions[row[0].row].height = 18
    # Green header
    for cell in ws[1]:
        if cell.value:
            cell.fill = green_fill

def reorder_columns(ws, desired_order):
    headers = [cell.value for cell in ws[1]]
    headers_lower = [str(h).lower() if h else "" for h in headers]
    desired_order_lower = [h.lower() for h in desired_order]

    # Find indices for desired columns
    ordered_indices = []
    for col in desired_order_lower:
        if col in headers_lower:
            ordered_indices.append(headers_lower.index(col))
    # Add remaining columns sorted by name
    remaining = sorted([i for i, h in enumerate(headers) if i not in ordered_indices], key=lambda x: str(headers[x]).lower())
    ordered_indices += remaining

    # Build new rows with columns in the desired order
    all_rows = list(ws.iter_rows(values_only=True))
    ws.delete_rows(1, ws.max_row)
    for row in all_rows:
        new_row = [row[i] if i < len(row) else "" for i in ordered_indices]
        ws.append(new_row)

def enrich_upsert_sheet_with_dictionary():
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    UPSET_FILE = os.path.join(SCRIPT_DIR, "SF New Hire API UpsertV1.xlsx")
    DICT_FILE = os.path.join(SCRIPT_DIR, "EC_APIField_Metadata.xlsx")

    wb_upsert = openpyxl.load_workbook(UPSET_FILE)
    ws_upsert = wb_upsert["Upsert API Field Attribute"]

    wb_dict = openpyxl.load_workbook(DICT_FILE, data_only=True)
    ws_dict = wb_dict["Simple EC Data API Dictionary"]

    dict_headers = [cell.value for cell in ws_dict[1]]
    lookup_cols = ["label","Type", "Key", "required", "picklist", "MaxLength", "NavigationField", "visible", "filterable", "sortable", "upsertable","creatable", "updatable"]
    entity_idx = dict_headers.index("Entity")
    name_idx = dict_headers.index("Name")
    dict_headers_lower = [str(h).lower() if h else "" for h in dict_headers]
    col_indices = [dict_headers_lower.index(col.lower()) for col in lookup_cols]

    lookup = {}
    for row in ws_dict.iter_rows(min_row=2, values_only=True):
        key = (str(row[entity_idx]), str(row[name_idx]))
        values = [row[i] for i in col_indices]
        lookup[key] = values

    # Add new columns to Upsert API Field Attribute sheet if not already present
    upsert_headers = [cell.value for cell in ws_upsert[1]]
    for i, col in enumerate(lookup_cols, start=3):  # after Entity, Field
        ws_upsert.cell(row=1, column=i+1, value=col)

    # Fill in the matched values
    for row in ws_upsert.iter_rows(min_row=2):
        entity = str(row[0].value)
        field = str(row[1].value)
        values = lookup.get((entity, field), [""] * len(lookup_cols))
        for i, val in enumerate(values, start=3):
            row[i].value = val

    # Reorder columns as required
    # desired_order = ["Entity", "Field", "Label","Type", "Key", "Required", "Picklist", "MaxLength", "Sample Value","Creatable", "Updatable", "NavigationField", "Visible", "Filterable", "Sortable", "Upsertable"]
    # reorder_columns(ws_upsert, desired_order)
    capitalize_headers(ws_upsert)
    autofit_and_style(ws_upsert)
    wb_upsert.save(UPSET_FILE)
    print("Upsert API Field Attribute sheet enriched with metadata columns.")

def enrich_api_entity_sheet():
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    UPSET_FILE = os.path.join(SCRIPT_DIR, "SF New Hire API UpsertV1.xlsx")
    NEW_FILE = os.path.join(SCRIPT_DIR, "New Hire API DocumentV1.xlsx")
    EC_METADATA_FILE = os.path.join(SCRIPT_DIR, "EC_APIField_Metadata.xlsx")
    ATTR_FILE = os.path.join(SCRIPT_DIR, "Employee Central API AttributeV2.xlsx")

    wb_upsert = openpyxl.load_workbook(UPSET_FILE)
    ws_api_entity = wb_upsert["API Entity"]

    wb_ecmeta = openpyxl.load_workbook(EC_METADATA_FILE, data_only=True)
    ws_ec_entity = wb_ecmeta["EC Entity"]

    wb_attr = openpyxl.load_workbook(ATTR_FILE, data_only=True)
    ws_person_emp = wb_attr["Person+Employment"]

    # --- Build lookups ---
    attr_headers = [cell.value for cell in ws_person_emp[1]]
    attr_headers_lower = [str(h).lower() if h else "" for h in attr_headers]
    attr_cols = ["Introduction", "BusinessKeys", "Effective-Date", "PersonEntityElement"]
    attr_indices = [attr_headers_lower.index(col.lower()) for col in attr_cols]
    attr_entity_idx = attr_headers_lower.index("entity")

    # Build a lookup for entity -> values
    attr_lookup = {}
    for row in ws_person_emp.iter_rows(min_row=2, values_only=True):
        key = str(row[attr_entity_idx]).lower()  # Lowercase for case-insensitive match
        values = [row[i] for i in attr_indices]
        attr_lookup[key] = values

    # --- Add missing columns to API Entity sheet ---
    api_headers = [cell.value for cell in ws_api_entity[1]]
    api_headers_lower = [str(h).lower() if h else "" for h in api_headers]
    # Only add columns if not already present
    for col in attr_cols:
        if col.lower() not in api_headers_lower:
            ws_api_entity.cell(row=1, column=ws_api_entity.max_column + 1, value=col)

    # Refresh headers after adding
    api_headers = [cell.value for cell in ws_api_entity[1]]
    api_headers_lower = [str(h).lower() if h else "" for h in api_headers]

    # --- Fill in matched values ---
    for row in ws_api_entity.iter_rows(min_row=2):
        api_entity = str(row[0].value).lower()  # Lowercase for case-insensitive match
        if api_entity in attr_lookup:
            values = attr_lookup[api_entity]
            for idx, val in enumerate(values):
                col_name = attr_cols[idx]
                if col_name.lower() in api_headers_lower:
                    col_idx = api_headers_lower.index(col_name.lower()) + 1  # openpyxl is 1-based
                    ws_api_entity.cell(row=row[0].row, column=col_idx, value=val)

    capitalize_headers(ws_api_entity)
    autofit_and_style(ws_api_entity)
    # Save to new file
    wb_upsert.save(NEW_FILE)
    print(f"API Entity sheet enriched and exported to {NEW_FILE}")

def clean_upsert_and_api_sample():
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    UPSET_FILE = os.path.join(SCRIPT_DIR, "New Hire API DocumentV1.xlsx")
    wb = openpyxl.load_workbook(UPSET_FILE)
    ws_upsert = wb["Upsert API Field Attribute"]
    ws_api_entity = wb["API Entity"]

    # Get header indices for upsert sheet
    upsert_headers = [cell.value for cell in ws_upsert[1]]
    upsert_headers_lower = [str(h).lower() if h else "" for h in upsert_headers]
    entity_idx = upsert_headers_lower.index("entity")
    field_idx = upsert_headers_lower.index("field")
    upsertable_idx = upsert_headers_lower.index("upsertable")
    sample_value_idx = upsert_headers_lower.index("sample value")  # Add this line

    # Collect rows to delete and keys to remove
    rows_to_delete = []
    keys_to_remove = {}  # {entity: set(fields)}

    for i, row in enumerate(ws_upsert.iter_rows(min_row=2), start=2):
        entity = str(row[entity_idx].value) if row[entity_idx].value else ""
        field = str(row[field_idx].value) if row[field_idx].value else ""
        upsertable = str(row[upsertable_idx].value) if row[upsertable_idx].value else ""
        sample_value = str(row[sample_value_idx].value) if row[sample_value_idx].value else ""
        # Updated Criteria
        if (not upsertable or upsertable.strip() == "false") and field != "__metadata":
            rows_to_delete.append(i)
            keys_to_remove.setdefault(entity, set()).add(field)
            continue
        # Criteria 2
        if field.lower() == "operation":
            rows_to_delete.append(i)
            keys_to_remove.setdefault(entity, set()).add(field)
            continue
        # Criteria 3
        if entity == "User" and field not in ["userId", "status", "username", "firstName", "lastName", "__metadata"]:
            rows_to_delete.append(i)
            keys_to_remove.setdefault(entity, set()).add(field)
            continue
        # Criteria 4: Blank sample value for PaymentInformationDetailV3
        if entity == "PaymentInformationDetailV3" and sample_value == "":
            rows_to_delete.append(i)
            keys_to_remove.setdefault(entity, set()).add(field)
            continue

    # Remove rows in reverse order to avoid index shifting
    for i in sorted(rows_to_delete, reverse=True):
        ws_upsert.delete_rows(i)

    # Update API Sample Upsert JSON in API Entity sheet
    api_headers = [cell.value for cell in ws_api_entity[1]]
    api_headers_lower = [str(h).lower() if h else "" for h in api_headers]
    entity_idx_api = api_headers_lower.index("entity")
    api_sample_idx = api_headers_lower.index("api sample upsert")

    for row in ws_api_entity.iter_rows(min_row=2):
        entity = str(row[entity_idx_api].value)
        if entity in keys_to_remove:
            json_str = row[api_sample_idx].value
            if not json_str:
                continue
            # Try to parse as JSON (handle single quotes if needed)
            try:
                # Try normal JSON
                data = json.loads(json_str)
            except Exception:
                # Try replacing single quotes with double quotes for parsing
                try:
                    data = json.loads(json_str.replace("'", '"'))
                except Exception:
                    continue  # skip if can't parse
            # Remove keys
            for key in keys_to_remove[entity]:
                if key in data:
                    del data[key]
            # Write back as JSON string (with ensure_ascii=False for non-ASCII)
            row[api_sample_idx].value = json.dumps(data, ensure_ascii=False)

    wb.save(UPSET_FILE)
    print("Redundant rows and keys removed, and API Sample Upsert updated.")

def main():
    enrich_upsert_sheet_with_dictionary()
    enrich_api_entity_sheet()
    clean_upsert_and_api_sample()

if __name__ == "__main__":
    main()